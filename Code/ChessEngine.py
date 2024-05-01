#importing libries
import os
from openpyxl import load_workbook
import random
import chess

#opening database
os.chdir('C:\\Users\\ejink\\OneDrive\\Desktop\\uni\\Y3\\2-4.5-finalproject\\Code')
fullcm = load_workbook('ChessMemory.xlsx')
cm = fullcm['ChessMemory']

def setUp():
    return chess.Board()

def guiCom(game):
    list=[]
    while True:
        line = input()
        if line == 'uci':
            print('id name ChessEngine\n')
            print('id author Enoch Jozue Krzok\n')
            print('uciok')
        elif line == 'quit':
            return list, False
        elif line == 'isready':
            print('readyok')
        elif line.startswith("ucinewgame"):
            game.reset()
        elif line.startswith("position"):
            if len(line) > len('position startpos '):
                game.push(chess.Move.from_uci(line[line.rindex(' ')+1:]))
        elif line.startswith('go'):
            game=posMove(game)
            list.append(remMove(game))
        elif line.startswith("result"):
            result = line.split(" ")[1]
            print("Game result:", result)
            dataUp(list)
    
                     
def posMove(game):
    list=[]
    for x in game.legal_moves:
        list.append(x)
    no = random.choice(list)
    game.push(no)
    print('bestmove',no)
    return game

def remMove(game):
    bFEN=game.board_fen()
    cFEN=game.castling_xfen()
    eFEN=game.epd().split()[3]
    fen = f'{bFEN} {cFEN} {eFEN}'
    return fen

def dataUp(list):
    list1=[]
    colFEN = 1
    for i in list:
        fen=i
        for n in cm.iter_rows(min_row=2, values_only=True):
            cellVal=n[fen-1]
            if cellVal == fen:
                print(f"Found '{fen}' in column {colFEN}, Row {cm.max_row - len(cm._cells) // cm.max_column + 1}")
                break
        else:
            print('not found')

    
    cm.save("ChessMemory.xlsx")

if __name__ == '__main__':
    game = setUp()
    list =guiCom(game)
